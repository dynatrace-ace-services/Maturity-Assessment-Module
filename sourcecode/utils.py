from functools import wraps
import time
import pandas as pd
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.table import Table

def exec_time(f):
    @wraps(f)
    def wrapper(*args, **kwargs):
        start_time = time.time()
        ret_func = f(*args, **kwargs)
        time_spent = time.time() - start_time
        args[0].logger.info("Execution of the code block took {:.2f} seconds".format(time_spent))
        return ret_func
    return wrapper

def exception_catcher(f):
    @wraps(f)
    def wrapper(*args, **kwargs):
        try:
            return f(*args, **kwargs)
        except Exception as exc:
            args[0].logger.exception("Exception caught at {}.{} : {}".format(f,__name__,exc))
            #raise
    return wrapper

def Excel_Exporter(filename, data, sheet):
    df_writer = pd.ExcelWriter('{}'.format(filename), engine = 'openpyxl', mode='a', if_sheet_exists='replace')
    data.to_excel(df_writer, sheet_name=sheet, index=False)
    ws = df_writer.sheets[sheet]
    #table = Table(displayName="Data", ref="A1:" + get_column_letter(ws.max_column) + str(ws.max_row))
    #ws.add_table(table)
    df_writer.close()
    